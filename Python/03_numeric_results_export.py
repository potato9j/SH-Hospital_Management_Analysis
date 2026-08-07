# -*- coding: utf-8 -*-
"""
Export a single submission-ready numeric results workbook and Markdown summary.

This script consolidates the values used in the thesis: financial ratios,
technical statistics, correlations, OLS/HC3/cluster SE, VIF, Cook's distance,
leverage, sensitivity analysis, RE/TWFE and Hausman tests.
"""
from __future__ import annotations

from pathlib import Path
import shutil
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from importlib.machinery import SourceFileLoader
cfg = SourceFileLoader("cfg", str(Path(__file__).with_name("00_local_paths.py"))).load_module()


def copy_sheet_values(src_ws, dst_ws) -> None:
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.number_format = cell.number_format
    dst_ws.freeze_panes = "A2"


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    side = Side(style="thin", color="D9D9D9")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=side)
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, min(ws.max_column, 20) + 1):
        letter = get_column_letter(col)
        width = min(38, max(10, max((len(str(ws.cell(r, col).value)) if ws.cell(r, col).value is not None else 0) for r in range(1, min(ws.max_row, 80) + 1)) + 2))
        ws.column_dimensions[letter].width = width


def add_readme_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("README_INDEX", 0)
    rows = [
        ["항목", "내용"],
        ["파일명", "hospital_management_full_numeric_results.xlsx"],
        ["목적", "논문 제출용 전체 수치결과 자료"],
        ["분석단위", "병원-환자경험평가연도"],
        ["관측치", "47개 상급종합병원 × 2021년·2023년 = 94개"],
        ["주 종속변수", "patient_experience_index"],
        ["주 의료질 변수", "year_matched_quality_index"],
        ["주 재무성과 변수", "main_medical_income_margin_pct"],
        ["주 비용구조 변수", "main_labor_cost_ratio_pct"],
        ["표준오차", "본문 기준은 병원 단위 군집표준오차, HC3와 OLS는 비교용"],
        ["패널모형", "Random Effects, Two-Way Fixed Effects, Hausman Test 포함"],
    ]
    for r, row in enumerate(rows, 1):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
    style_sheet(ws)


def add_variable_map(wb: Workbook) -> None:
    ws = wb.create_sheet("Variable_Name_Map", 1)
    rows = [
        ["논문 명칭", "분석용 변수명", "역할"],
        ["환자경험지수", "patient_experience_index", "종속변수"],
        ["간호사 영역", "nurse", "하위 종속변수"],
        ["의사 영역", "doctor", "하위 종속변수"],
        ["투약 및 치료과정", "treatment", "하위 종속변수"],
        ["병원환경", "environment", "하위 종속변수"],
        ["환자권리보장", "rights", "하위 종속변수"],
        ["전반적 평가", "overall", "하위 종속변수"],
        ["의료수익의료이익률", "main_medical_income_margin_pct", "핵심 독립변수"],
        ["인건비율", "main_labor_cost_ratio_pct", "핵심 독립변수"],
        ["평가연도 매칭 의료질지수", "year_matched_quality_index", "주요 설명변수"],
        ["최신가용 의료질지수", "latest_quality_index", "보조 민감도 변수"],
        ["로그 병상수", "log_total_beds", "통제변수"],
        ["100병상당 의사수", "doctors_per_100_beds", "통제변수"],
        ["간호등급", "nursing_grade_main_num", "통제변수"],
        ["100병상당 의료장비수", "equipment_per_100_beds", "통제변수"],
        ["수도권 여부", "metro_area", "통제변수"],
        ["2023년 더미", "year_2023_dummy", "통제변수"],
    ]
    for r, row in enumerate(rows, 1):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
    style_sheet(ws)


def consolidate_workbooks() -> Path:
    cfg.RESULT_DIR.mkdir(parents=True, exist_ok=True)
    source_main = cfg.PROJECT_DIR / "outputs" / "year_matched_full_revision_outputs" / "year_matched_full_recalculation_results.xlsx"
    if not source_main.exists():
        # Fallback for the current project layout used during development.
        source_main = Path("/mnt/data/year_matched_full_revision_outputs/year_matched_full_recalculation_results.xlsx")
    wb_src = load_workbook(source_main, data_only=True)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    add_readme_sheet(wb_out)
    add_variable_map(wb_out)

    for sheet_name in wb_src.sheetnames:
        dst_name = sheet_name[:31]
        ws_dst = wb_out.create_sheet(dst_name)
        copy_sheet_values(wb_src[sheet_name], ws_dst)
        style_sheet(ws_dst)

    # Add finance ratio stage sheets if available.
    finance_stage = cfg.STAGE_FILES.get("finance_ratios")
    if finance_stage and finance_stage.exists():
        fin_wb = load_workbook(finance_stage, data_only=True)
        for sheet_name in ["Summary", "Financial_Ratios", "Ratio_Definitions", "Data_Quality"]:
            if sheet_name in fin_wb.sheetnames:
                ws_dst = wb_out.create_sheet(("FIN_" + sheet_name)[:31])
                copy_sheet_values(fin_wb[sheet_name], ws_dst)
                style_sheet(ws_dst)

    wb_out.save(cfg.FINAL_RESULTS_XLSX)
    return cfg.FINAL_RESULTS_XLSX


def write_markdown_summary(xlsx_path: Path) -> Path:
    md = cfg.FINAL_RESULTS_MD
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_lines = "\n".join([f"- `{s}`" for s in wb.sheetnames])
    text = f"""# 병원경영분석 전체 수치결과 자료

이 파일은 논문 제출 시 함께 제출할 수 있도록 구성한 전체 수치결과 자료의 색인입니다.

## 핵심 분석 구조

- 분석대상: 제5기 상급종합병원 47개소
- 분석단위: 병원-환자경험평가연도
- 관측치: 94개
- 주 종속변수: `patient_experience_index`
- 주 의료질 변수: `year_matched_quality_index`
- 주 회귀모형: 병원 단위 군집표준오차 기준 통합모형
- 보조모형: HC3, 일반 OLS, 민감도 분석, Random Effects, Two-Way Fixed Effects, Hausman 검정

## 포함 시트

{sheet_lines}

## 제출 시 설명 문장 예시

본 연구는 재현성을 확보하기 위해 원자료 처리, 평가연도 매칭 의료질지수 산출, 재무비율 계산, 회귀분석, 표준오차 보정, 회귀진단, 민감도 분석, 패널모형 분석 결과를 하나의 수치결과 파일로 정리하였다. 분석용 변수명은 본문과 부록의 변수정의표에 맞추었다.
"""
    md.write_text(text, encoding="utf-8")
    return md


def main() -> None:
    out = consolidate_workbooks()
    md = write_markdown_summary(out)
    print(f"[DONE] Results workbook: {out}")
    print(f"[DONE] Markdown summary: {md}")


if __name__ == "__main__":
    main()
